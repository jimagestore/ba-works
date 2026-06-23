"""Convert BA spec xlsx files to structured markdown."""
import openpyxl, os, glob
from openpyxl.cell.cell import MergedCell

FIELD_NAME_COL = 2  # B
TYPE_COL = 3        # C
REQUIRED_COL = 4    # D
DESC_COL = 5        # E
DEV_COL = 9         # I
TEST_COL = 10       # J

def safe_str(v):
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        v = int(v)
    return str(v)

def read_row_vals(ws, row_num):
    vals = {}
    for c in ws[row_num]:
        if c.value is not None:
            vals[c.column] = c.value
    return vals

def has_merged(ws, row_num, col):
    for m in ws.merged_cells.ranges:
        if m.min_row <= row_num <= m.max_row and m.min_col <= col <= m.max_col:
            return True
    return False

def get_merge_min(ws, row_num, minc=2, maxc=8):
    """Return the minimum column that is merged within [minc, maxc]."""
    merged = set()
    for m in ws.merged_cells.ranges:
        if m.min_row <= row_num <= m.max_row:
            for ci in range(m.min_col, m.max_col + 1):
                if minc <= ci <= maxc:
                    merged.add(ci)
    return min(merged) if merged else None

def convert_xlsx_to_md(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # --- Find Field Name row ---
    field_row = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=10, values_only=False):
        for c in row:
            if c.value and isinstance(c.value, str) and c.value.strip() == "Field Name":
                field_row = c.row
                break
        if field_row:
            break
    if not field_row:
        return f"# {os.path.splitext(os.path.basename(xlsx_path))[0]}\n\n*No Field Name row found*"

    lines = []
    title = os.path.splitext(os.path.basename(xlsx_path))[0]
    lines.append(f"# {title}")
    lines.append("")

    # ========== HEADER SECTION ==========
    header_data = {
        'story_id': '', 'conversation': '', 'design': '', 'cards': [],
        'precondition_path': ''
    }
    current_section = None
    for r in range(1, field_row):
        vals = read_row_vals(ws, r)
        if not vals:
            continue
        a = safe_str(vals.get(1, ''))
        b = safe_str(vals.get(2, ''))

        if a == 'User Story':
            header_data['story_id'] = b
            current_section = 'story'
        elif a == 'Conversation':
            header_data['conversation'] = b
            current_section = 'conversation'
        elif a == 'Design':
            header_data['design'] = b
            current_section = 'design'
        elif a == 'Card':
            header_data['cards'].append(b)
            current_section = 'card'
        elif a == 'Confirmation':
            current_section = 'confirmation'
        elif a == '' and b:
            if current_section == 'card':
                header_data['cards'].append(b)
        if b and ('Pre-condition' in b or 'pre-condition' in b.lower()):
            header_data['precondition_path'] = b

    lines.append("---")
    lines.append("")
    if header_data['story_id']:
        lines.append(f"**User story:** {header_data['story_id']}")
        lines.append("")
    if header_data['conversation']:
        conv = header_data['conversation']
        if conv.startswith('='):
            conv = "*VLOOKUP: see Agenda sheet*"
        lines.append(f"**Conversation:** {conv}")
        lines.append("")
    if header_data['design']:
        lines.append(f"**Design:** {header_data['design']}")
        lines.append("")
    if header_data['cards']:
        lines.append("**Card:**")
        for c in header_data['cards']:
            lines.append(f"- {c}")
        lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("**Confirmation:**")
    lines.append("")
    if header_data['precondition_path']:
        lines.append("**Pre-condition / Path:**")
        for line in header_data['precondition_path'].split('\n'):
            lines.append(line)
        lines.append("")
    lines.append("---")
    lines.append("")

    # ========== TABLE SECTION ==========
    items = []

    for r in range(field_row + 1, ws.max_row + 1):
        vals = read_row_vals(ws, r)
        if not vals:
            continue

        b_val = safe_str(vals.get(FIELD_NAME_COL, ''))
        c_val = safe_str(vals.get(TYPE_COL, ''))
        d_val = safe_str(vals.get(REQUIRED_COL, ''))
        e_val = safe_str(vals.get(DESC_COL, ''))
        dev_val = safe_str(vals.get(DEV_COL, ''))
        test_val = safe_str(vals.get(TEST_COL, ''))

        has_b = bool(b_val)
        has_type = bool(c_val)
        has_req = bool(d_val)
        has_desc = bool(e_val)

        # Determine merge pattern
        b_merged = has_merged(ws, r, FIELD_NAME_COL)
        c_merged = has_merged(ws, r, TYPE_COL)
        merge_min = get_merge_min(ws, r)

        # Classify row type based on merge pattern
        if b_merged and b_val:
            # ALL merged (B-H) = section name
            items.append({'type': 'section', 'name': b_val})

        elif merge_min == 3 and has_b and not (has_type or has_req or has_desc):
            # C-H merged, B separate = BR row
            items.append({
                'type': 'br', 'name': b_val, 'description': '',
                'dev': dev_val, 'test': test_val
            })

        elif merge_min == 3 and not has_b and c_val:
            # C-H merged, B empty = continuation note for BR
            if items:
                items.append({
                    'type': 'note', 'text': c_val,
                    'dev': dev_val, 'test': test_val
                })

        elif merge_min == 3 and has_b:
            # C-H merged, B has value = BR (e.g. "BR1" with description in C)
            items.append({
                'type': 'br', 'name': b_val, 'description': c_val,
                'dev': dev_val, 'test': test_val
            })

        elif merge_min == 5 and has_b and (has_type or has_req or has_desc):
            # E-H merged, B/C/D separate = UI Element
            items.append({
                'type': 'element', 'name': b_val,
                'element_type': c_val, 'required': d_val,
                'description': e_val,
                'dev': dev_val, 'test': test_val
            })

        elif merge_min == 5 and has_b and not has_type and not has_req and not has_desc:
            # E-H merged, B has value but no type/req/desc
            items.append({
                'type': 'element', 'name': b_val,
                'element_type': '', 'required': '', 'description': '',
                'dev': dev_val, 'test': test_val
            })

        elif merge_min == 5 and not has_b and (dev_val or test_val):
            # DEV/Test only row
            parts = []
            if dev_val: parts.append(f"DEV: {dev_val}")
            if test_val: parts.append(f"Test: {test_val}")
            if items:
                items.append({'type': 'note', 'text': ' | '.join(parts)})

        elif not b_merged and not c_merged and has_b:
            # No special merge, but has content = generic element
            items.append({
                'type': 'element', 'name': b_val,
                'element_type': c_val, 'required': d_val,
                'description': e_val or c_val,
                'dev': dev_val, 'test': test_val
            })

    # Render
    lines.append("## Requirements")
    lines.append("")

    for item in items:
        if item['type'] == 'section':
            lines.append(f"### {item['name']}")
            lines.append("")

        elif item['type'] == 'element':
            lines.append(f"- **{item['name']}**")
            if item['element_type']:
                lines.append(f"  - **Type:** {item['element_type']}")
            if item['required']:
                lines.append(f"  - **Required:** {item['required']}")
            if item['description']:
                desc = item['description'].replace('\n', '\n  ')
                lines.append(f"  - **Description:** {desc}")
            devtest = []
            if item['dev']: devtest.append(f"DEV: {item['dev']}")
            if item['test']: devtest.append(f"Test: {item['test']}")
            if devtest:
                lines.append(f"  - *{' | '.join(devtest)}*")
            lines.append("")

        elif item['type'] == 'br':
            lines.append(f"- **{item['name']}**")
            if item['description']:
                desc = item['description'].replace('\n', '\n  ')
                lines.append(f"  - {desc}")
            devtest = []
            if item['dev']: devtest.append(f"DEV: {item['dev']}")
            if item['test']: devtest.append(f"Test: {item['test']}")
            if devtest:
                lines.append(f"  - *{' | '.join(devtest)}*")
            lines.append("")

        elif item['type'] == 'note':
            lines.append(f"  - *{item['text']}*")
            lines.append("")

    return '\n'.join(lines)


# Main: convert all xlsx files
src_dir = os.path.dirname(os.path.abspath(__file__))
for f in sorted(glob.glob(os.path.join(src_dir, "*.xlsx"))):
    md = convert_xlsx_to_md(f)
    md_path = os.path.splitext(f)[0] + ".md"
    with open(md_path, "w", encoding="utf-8") as out:
        out.write(md)
    print(f"OK {os.path.basename(f)} -> {os.path.basename(md_path)} ({len(md)} chars)")
